#!/usr/bin/env python3
"""
augment_existing_dataset.py

Takes an existing stego_training.xlsx dataset and creates augmented variants
of the images, adding them to a new Excel file.

This is useful if you already have a dataset and want to expand it with augmented versions
without regenerating everything.
"""

import pandas as pd
from pathlib import Path
import logging
from tqdm import tqdm
from image_augment import ImageAugmentor
from openpyxl import Workbook
import random

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ============ CONFIGURATION ============
INPUT_EXCEL = 'dataGen/stego_training.xlsx'
OUTPUT_EXCEL = 'dataGen/stego_training_augmented.xlsx'
OUTPUT_DIR = 'dataGen/sten_data_augmented'

AUGMENT_CLEAN = True       # Augment clean images?
AUGMENT_STEGO = True       # Augment stego images?
NUM_VARIANTS = 2           # How many augmented versions per image
AUGMENT_RATIO = 0.5        # Fraction of images to augment (0.5 = 50%)
USE_MILD_AUGMENTATION = True  # Use mild parameters to preserve stego data
RANDOM_SEED = 42
# =======================================


def main():
    """Augment existing dataset with image transformations."""
    input_path = Path(INPUT_EXCEL)
    output_path = Path(OUTPUT_EXCEL)
    output_dir = Path(OUTPUT_DIR)
    
    output_dir.mkdir(parents=True, exist_ok=True)
    random.seed(RANDOM_SEED)
    
    # Load existing dataset
    logger.info("Loading dataset from %s", input_path)
    df = pd.read_excel(input_path)
    
    # Expected columns
    if 'File Path' not in df.columns or 'Stegnography Applied?' not in df.columns:
        raise ValueError("Excel must have 'File Path' and 'Stegnography Applied?' columns")
    
    logger.info("Loaded %d images from dataset", len(df))
    logger.info("  Clean images: %d", sum(~df['Stegnography Applied?']))
    logger.info("  Stego images: %d", sum(df['Stegnography Applied?']))
    
    # Initialize augmentor
    augmentor = ImageAugmentor(seed=RANDOM_SEED)
    
    # Create new workbook with original data + augmented
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "File Path"
    ws["B1"] = "Stegnography Applied?"
    row_count = 2
    
    # Copy all original entries
    for _, row in df.iterrows():
        ws.cell(row=row_count, column=1, value=row['File Path'])
        ws.cell(row=row_count, column=2, value=row['Stegnography Applied?'])
        row_count += 1
    
    original_count = len(df)
    augmented_count = 0
    
    # Process each image
    for idx, row in tqdm(df.iterrows(), total=len(df), desc='Augmenting images'):
        file_path = Path(row['File Path'])
        is_stego = row['Stegnography Applied?']
        
        # Decide whether to augment this image
        should_augment = False
        if is_stego and AUGMENT_STEGO:
            should_augment = random.random() < AUGMENT_RATIO
        elif not is_stego and AUGMENT_CLEAN:
            should_augment = random.random() < AUGMENT_RATIO
        
        if not should_augment:
            continue
        
        # Check if source file exists
        if not file_path.exists():
            logger.warning("Source file not found: %s", file_path)
            continue
        
        # Create augmented variants
        for var_idx in range(NUM_VARIANTS):
            aug_name = f"aug_{var_idx}_{file_path.name}"
            aug_path = output_dir / aug_name
            
            try:
                success = augmentor.augment_and_save(
                    file_path,
                    aug_path,
                    transforms='random',
                    mild=USE_MILD_AUGMENTATION
                )
                
                if success and aug_path.exists():
                    # Add to Excel
                    ws.cell(row=row_count, column=1, value=str(aug_path.resolve()))
                    ws.cell(row=row_count, column=2, value=is_stego)
                    row_count += 1
                    augmented_count += 1
            except Exception as e:
                logger.warning("Failed to augment %s (variant %d): %s", file_path, var_idx, e)
    
    # Save augmented dataset
    wb.save(output_path)
    
    logger.info("=" * 60)
    logger.info("Augmentation complete!")
    logger.info("  Original images: %d", original_count)
    logger.info("  Augmented variants created: %d", augmented_count)
    logger.info("  Total dataset size: %d", row_count - 2)
    logger.info("  Output Excel: %s", output_path)
    logger.info("  Augmented images dir: %s", output_dir)
    logger.info("=" * 60)


if __name__ == '__main__':
    main()
