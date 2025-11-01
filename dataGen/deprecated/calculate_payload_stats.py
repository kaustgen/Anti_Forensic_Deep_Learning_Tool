#!/usr/bin/env python3
# Author: Kaleb Austgen
# Date: 10/22/25
# Purpose: Payload Statistics
"""
calculate_payload_stats.py

Analyzes stego images from the Excel workbook and calculates payload statistics.
For each stego image, finds its original cover, computes payload size, and 
calculates bits per non-zero AC DCT coefficient (the most relevant metric for 
steghide-based steganography).

Adds a new column "Payload (bpp AC DCT)" to the Excel workbook.
"""

import logging
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from PIL import Image
import numpy as np
import re

# Try to import jpegio for DCT coefficient extraction
try:
    import jpegio as jio
    HAVE_JPEGLIB = True
except ImportError:
    jio = None
    HAVE_JPEGLIB = False
    print("Warning: jpegio not available. Install with: pip install jpegio")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def get_file_size_bytes(filepath):
    """Get file size in bytes."""
    try:
        return filepath.stat().st_size
    except Exception as e:
        logger.error("Failed to get size for %s: %s", filepath, e)
        return None


def count_nonzero_ac_coefficients(image_path):
    """
    Count the number of non-zero AC DCT coefficients in a JPEG image.
    
    For steghide, this is the most relevant capacity metric since it embeds
    data by replacing LSBs of non-zero AC coefficients.
    
    Returns:
        int: Total count of non-zero AC coefficients across all channels
    """
    if not HAVE_JPEGLIB:
        logger.warning("Cannot count DCT coefficients without jpegio")
        return None
    
    try:
        jpeg_struct = jio.read(str(image_path))
        coef_arrays = jpeg_struct.coef_arrays
        
        if coef_arrays is None or len(coef_arrays) == 0:
            logger.warning("No DCT coefficients found in %s", image_path)
            return None
        
        total_nonzero_ac = 0
        
        for channel_idx, coeffs in enumerate(coef_arrays):
            # Reshape into 8x8 blocks
            h, w = coeffs.shape
            h_blocks = h // 8
            w_blocks = w // 8
            
            # Trim to multiple of 8
            coeffs_trimmed = coeffs[:h_blocks*8, :w_blocks*8]
            
            # Reshape to blocks: (num_blocks, 8, 8)
            blocks = coeffs_trimmed.reshape(h_blocks, 8, w_blocks, 8).transpose(0, 2, 1, 3)
            blocks = blocks.reshape(-1, 8, 8)
            
            # For each block, count non-zero AC coefficients
            # DC coefficient is at [0, 0], so we exclude it
            for block in blocks:
                # Flatten block and skip DC (first element)
                ac_coeffs = block.flatten()[1:]  # Skip DC at index 0
                nonzero_ac = np.count_nonzero(ac_coeffs)
                total_nonzero_ac += nonzero_ac
        
        return total_nonzero_ac
    
    except Exception as e:
        logger.error("Failed to count DCT coefficients for %s: %s", image_path, e)
        return None


def get_image_dimensions(image_path):
    """Get image dimensions (width, height)."""
    try:
        with Image.open(image_path) as img:
            return img.size  # (width, height)
    except Exception as e:
        logger.error("Failed to get dimensions for %s: %s", image_path, e)
        return None


def find_original_cover(stego_path, clean_images_dir):
    """
    Find the original cover image for a stego image.
    
    Strategy:
    1. If stego is in sten_data/, look for original in clean_images/
    2. If stego has "aug_" prefix, strip it and look for base stego
    3. Extract base name and find matching cover
    """
    stego_path = Path(stego_path)
    filename = stego_path.name
    
    # Case 1: Augmented stego image (aug_X_stego_Y.jpg)
    if 'aug_' in stego_path.name:
        # Extract the base stego name: aug_0_stego_photo-123.jpg -> stego_photo-123.jpg
        parts = stego_path.name.split('_')
        if 'stego' in parts:
            stego_idx = parts.index('stego')
            base_name = '_'.join(parts[stego_idx:])
            base_stego = stego_path.parent.parent / 'sten_data' / base_name
            
            if base_stego.exists():
                # Now find the cover for this base stego
                return find_original_cover(base_stego, clean_images_dir)
    
    # Case 2: Base stego image (stego_photo-123.jpg or stego_photo-123_0.jpg)
    if stego_path.name.startswith('stego_'):
        # Extract cover name: stego_photo-123_0.jpg -> photo-123.jpg
        base_name = stego_path.name.replace('stego_', '')
        
        # Remove trailing _N if present (stego_photo_0 -> photo)
        if '_' in base_name:
            parts = base_name.rsplit('_', 1)
            if len(parts) == 2 and parts[1].split('.')[0].isdigit():
                base_name = parts[0] + Path(base_name).suffix
        
        # Look for cover in clean_images directory
        for ext in ['.jpg', '.jpeg', '.png', '.JPG', '.JPEG', '.PNG']:
            cover_name = Path(base_name).stem + ext
            cover_path = clean_images_dir / cover_name
            if cover_path.exists():
                return cover_path
    
    # Case 3: Augmented clean image (aug_clean_X_photo-123.jpg)
    if 'aug_clean_' in stego_path.name:
        # Extract original name: aug_clean_0_photo-123.jpg -> photo-123.jpg
        parts = stego_path.name.split('_')
        if len(parts) >= 3:
            base_name = '_'.join(parts[3:])  # Skip 'aug', 'clean', 'X'
            cover_path = clean_images_dir / base_name
            if cover_path.exists():
                return cover_path
    
    # THE BELOW CODE IS FOR THE BOSS DATA SPECIFICALLY, YOU HAVE TO UPDATE PATHS MANUALLY
        # Pattern 1: aug_X_stego_NUMBER_Y.jpg → extract NUMBER
    match = re.search(r'aug_\d+_stego_(\d+)_\d+\.jpg', filename)
    if match:
        number = match.group(1)
        cover_name = f"{number}.jpg"
        cover_path = clean_images_dir / cover_name
        if cover_path.exists():
            return cover_path
    
    # Pattern 2: stego_NUMBER_Y.jpg → extract NUMBER
    match = re.search(r'stego_(\d+)_\d+\.jpg', filename)
    if match:
        number = match.group(1)
        cover_name = f"{number}.jpg"
        cover_path = clean_images_dir / cover_name
        if cover_path.exists():
            return cover_path
    
    # Pattern 3: Direct number (clean images): NUMBER.jpg
    match = re.search(r'^(\d+)\.jpg$', filename)
    if match:
        # This is already a clean image, return itself
        return stego_path
    
    logger.warning("Could not find original cover for: %s", stego_path)
    return None


def calculate_payload_metrics(stego_path, cover_path):
    """
    Calculate payload statistics for a stego image.
    
    Returns:
        dict with keys:
            - payload_bytes: Raw payload size in bytes (file size difference)
            - payload_bits: Payload in bits
            - dimensions: (width, height)
            - total_pixels: width * height
            - bpp: Bits per pixel
            - nonzero_ac_coeffs: Count of non-zero AC DCT coefficients
            - bpp_ac_dct: Bits per non-zero AC DCT coefficient (most relevant)
    """
    stego_path = Path(stego_path)
    cover_path = Path(cover_path)
    
    # Get file sizes
    stego_size = get_file_size_bytes(stego_path)
    cover_size = get_file_size_bytes(cover_path)
    
    if stego_size is None or cover_size is None:
        return None
    
    # Calculate payload (note: this is approximate due to JPEG recompression)
    payload_bytes = stego_size - cover_size
    payload_bits = payload_bytes * 8
    
    # Get dimensions
    dimensions = get_image_dimensions(cover_path)
    if dimensions is None:
        return None
    
    width, height = dimensions
    total_pixels = width * height
    bpp = payload_bits / total_pixels if total_pixels > 0 else 0
    
    # Count non-zero AC DCT coefficients (most relevant metric for steghide)
    nonzero_ac = count_nonzero_ac_coefficients(cover_path)
    bpp_ac_dct = payload_bits / nonzero_ac if nonzero_ac and nonzero_ac > 0 else None
    
    return {
        'payload_bytes': payload_bytes,
        'payload_bits': payload_bits,
        'dimensions': dimensions,
        'total_pixels': total_pixels,
        'bpp': bpp,
        'nonzero_ac_coeffs': nonzero_ac,
        'bpp_ac_dct': bpp_ac_dct
    }


def main():
    """Main function to process Excel and add payload statistics."""
    
    # Configuration
    base_dir = Path(__file__).parent.parent
    excel_path = base_dir / 'BOSS_stego_training.xlsx'
    clean_images_dir = base_dir / 'BOSS1.01_Dataset'
    
    if not excel_path.exists():
        logger.error("Excel file not found: %s", excel_path)
        return
    
    if not clean_images_dir.exists():
        logger.error("Clean images directory not found: %s", clean_images_dir)
        return
    
    logger.info("Loading Excel workbook: %s", excel_path)
    
    # Load Excel with openpyxl to preserve formatting
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # Check if payload column already exists
    header_row = [cell.value for cell in ws[1]]
    if 'Payload (bpp AC DCT)' in header_row:
        logger.info("Payload column already exists, will update values")
        payload_col = header_row.index('Payload (bpp AC DCT)') + 1
    else:
        logger.info("Adding new column: Payload (bpp AC DCT)")
        payload_col = len(header_row) + 1
        ws.cell(row=1, column=payload_col, value='Payload (bpp AC DCT)')
    
    # Also add auxiliary columns for debugging/analysis
    aux_cols = {
        'Payload (bytes)': payload_col + 1,
        'Payload (bits)': payload_col + 2,
        'Image Dimensions': payload_col + 3,
        'Non-zero AC DCT': payload_col + 4,
    }
    
    for col_name, col_idx in aux_cols.items():
        if ws.cell(row=1, column=col_idx).value != col_name:
            ws.cell(row=1, column=col_idx, value=col_name)
    
    # Process each row
    total_rows = ws.max_row
    stego_count = 0
    clean_count = 0
    calculated_count = 0
    failed_count = 0
    
    logger.info("Processing %d rows...", total_rows - 1)
    
    for row_idx in range(2, total_rows + 1):  # Skip header
        file_path = ws.cell(row=row_idx, column=1).value
        is_stego = ws.cell(row=row_idx, column=2).value
        
        if not file_path:
            continue
        
        file_path = Path(file_path)
        
        # Convert string to boolean if needed
        if isinstance(is_stego, str):
            is_stego = is_stego.lower() in ['true', 'yes', '1']
        
        if is_stego:
            stego_count += 1
            
            # Find original cover
            cover_path = find_original_cover(file_path, clean_images_dir)
            
            if cover_path:
                # Calculate payload metrics
                metrics = calculate_payload_metrics(file_path, cover_path)
                
                if metrics:
                    # Write primary metric (bpp AC DCT)
                    if metrics['bpp_ac_dct'] is not None:
                        ws.cell(row=row_idx, column=payload_col, value=metrics['bpp_ac_dct'])
                    else:
                        ws.cell(row=row_idx, column=payload_col, value='N/A')
                    
                    # Write auxiliary data
                    ws.cell(row=row_idx, column=aux_cols['Payload (bytes)'], value=metrics['payload_bytes'])
                    ws.cell(row=row_idx, column=aux_cols['Payload (bits)'], value=metrics['payload_bits'])
                    ws.cell(row=row_idx, column=aux_cols['Image Dimensions'], 
                           value=f"{metrics['dimensions'][0]}×{metrics['dimensions'][1]}")
                    ws.cell(row=row_idx, column=aux_cols['Non-zero AC DCT'], 
                           value=metrics['nonzero_ac_coeffs'] if metrics['nonzero_ac_coeffs'] else 'N/A')
                    
                    calculated_count += 1
                else:
                    ws.cell(row=row_idx, column=payload_col, value='ERROR')
                    failed_count += 1
            else:
                ws.cell(row=row_idx, column=payload_col, value='NO COVER')
                failed_count += 1
        else:
            clean_count += 1
            # Clean images have no payload
            ws.cell(row=row_idx, column=payload_col, value='N/A')
    
    # Save workbook
    logger.info("Saving updated workbook...")
    wb.save(excel_path)
    
    # Print statistics
    logger.info("="*60)
    logger.info("PAYLOAD STATISTICS CALCULATION COMPLETE")
    logger.info("="*60)
    logger.info("  Total rows processed: %d", total_rows - 1)
    logger.info("  Clean images: %d", clean_count)
    logger.info("  Stego images: %d", stego_count)
    logger.info("  Successfully calculated: %d", calculated_count)
    logger.info("  Failed/missing: %d", failed_count)
    logger.info("="*60)
    logger.info("Updated Excel saved to: %s", excel_path)
    logger.info("")
    logger.info("New columns added:")
    logger.info("  - Payload (bpp AC DCT): Primary metric - bits per non-zero AC DCT coefficient")
    logger.info("  - Payload (bytes): Raw file size difference")
    logger.info("  - Payload (bits): Payload in bits")
    logger.info("  - Image Dimensions: Cover image width×height")
    logger.info("  - Non-zero AC DCT: Total non-zero AC coefficients in cover")
    logger.info("="*60)


if __name__ == '__main__':
    main()
