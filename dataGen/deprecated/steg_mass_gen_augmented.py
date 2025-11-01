#!/usr/bin/env python3
# Author: Kaleb Austgen
# Date: 10/21/25
# Purpose:
"""

Generates a large number of stego imgages as well as modified clean images 

Strategy:
1. For each clean cover image:
   - Add original clean image to dataset
   - Create N stego images with embedded secrets
   - Create augmented variants of stego images
2. Create augmented variants of clean images to balance the dataset
"""

from fpdf import FPDF
import random
from steg_gen import Steg_Gen
import logging
from pathlib import Path
from openpyxl import Workbook
from tqdm import tqdm
from image_augment import ImageAugmentor
import os
import tempfile
import random

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Configuration
# Configuration
base_dir = Path(__file__).parent
EXCEL_OUT = base_dir / 'BOSS_stego_training.xlsx'
PASSWORD = 'five'
TARGET_TOTAL = 10000  # Base number of STEGO images (clean will be generated to match)
AUGMENT_STEGO_RATIO = 0.0  # Total augmented = ratio × base (0.5 = 50% more augmented)
AUGMENT_CLEAN_RATIO = 0.0  # Percentage of clean images to augment
AUGMENT_MILD = True  # Use mild augmentations to preserve stego data

# Payload distribution configuration for balanced training
# BOSS images are 512x512 = 262,144 pixels
# Conservative max capacity: varies by image content (2-10KB typical)
PAYLOAD_DISTRIBUTION = {
    'low': {
        'ratio': 0.33,
        'secret_size_bytes': 512,  # 512 bytes - very safe for all images
        'description': 'Low payload (512B, ~0.015 bpp)'
    },
    'medium': {
        'ratio': 0.34,
        'secret_size_bytes': 2048,  # 2KB - safe for most images
        'description': 'Medium payload (2KB, ~0.06 bpp)'
    },
    'high': {
        'ratio': 0.33,
        'secret_size_bytes': 4096,  # 4KB - for high-complexity images only
        'description': 'High payload (4KB, ~0.12 bpp)'
    }
}


def generate_pdf(text, output_file):
    """Generate a simple PDF with given text."""
    output_file = Path(output_file)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt=text, ln=True, align="L")
    pdf.output(str(output_file))


# Sample sentences for PDFs
SENTENCES = [
        "This is a short note.",
        "Please review the attachment.",
        "System check complete.",
        "User reported no issues.",
        "Temporary token issued.",
        "Meeting at noon.",
        "Backup completed successfully.",
        "Refer to the log file.",
        "Operation completed.",
        "Low disk space warning.",
        "New device connected.",
        "Configuration updated.",
        "Restart required.",
        "Security scan passed.",
        "Pending approval needed.",
        "Contact support if needed.",
        "Last login was yesterday.",
        "Session expired - please re-authenticate.",
        "Username: admin_user",
        "Password: hunter2",
        "Email: user@example.com",
        "Temporary access granted.",
        "Database connection established.",
        "Query executed successfully.",
        "Transaction committed.",
        "Cache invalidated.",
        "Session token refreshed.",
        "API rate limit exceeded.",
        "Webhook delivered.",
        "Event published to queue.",
        "Message consumed from topic.",
        "Health check passed.",
        "Service running normally.",
        "Port 8080 listening.",
        "SSL certificate valid.",
        "CORS policy enforced.",
        "Request throttled.",
        "Retry attempted.",
        "Circuit breaker opened.",
        "Fallback invoked.",
        "Timeout set to 30s.",
        "Connection pool exhausted.",
        "Thread count increased.",
        "Memory usage: 75%",
        "CPU usage: 45%",
        "Disk I/O normal.",
        "Network latency: 12ms",
        "Packet loss: 0.01%",
        "Bandwidth: 100Mbps",
        "Load balancer active.",
        "Auto-scaling triggered.",
]


def estimate_image_capacity(img_path):
    """
    Estimate steghide embedding capacity for an image.
    Returns estimated capacity in bytes (conservative).
    """
    try:
        from PIL import Image
        img = Image.open(img_path)
        width, height = img.size
        
        # Conservative estimate: ~0.5 bits per pixel for steghide
        # (actual capacity depends on DCT coefficient distribution)
        pixels = width * height
        bits = pixels * 0.5
        bytes_capacity = int(bits / 8)
        
        # Add 50% safety margin (steghide needs more headroom than expected)
        return int(bytes_capacity * 0.5)
    except Exception:
        # Fallback: assume 512x512 image
        return 13000  # ~13KB conservative

# Generate secret data for the BOSS dataset
def generate_secret_data(size_bytes, method='random'):
    """
    Generate secret data of specified size.
    
    Args:
        size_bytes: Size of secret data in bytes
        method: 'random' for random bytes, 'text' for repeated lorem ipsum
    
    Returns:
        bytes: Secret data of exactly size_bytes length
    """
    if method == 'random':
        return os.urandom(size_bytes)
    elif method == 'text':
        # Generate lorem ipsum text of approximately size_bytes
        lorem = "Lorem ipsum dolor sit amet consectetur adipiscing elit " * 100
        text = (lorem * (size_bytes // len(lorem) + 1))[:size_bytes]
        return text.encode('utf-8')
    else:
        raise ValueError(f"Unknown method: {method}")


def main():
    # Find cover images
    base_dir = Path(__file__).parent
    images_dir = base_dir / 'BOSS1.01_Dataset'
    available_covers = []
    
    if images_dir.exists() and images_dir.is_dir():
        available_covers.extend(sorted(list(images_dir.glob('*.jpg')) + 
                                      list(images_dir.glob('*.jpeg')) + 
                                      list(images_dir.glob('*.png'))))
    
    if not available_covers:
        available_covers = sorted(list(base_dir.glob('*.jpg')) + 
                                 list(base_dir.glob('*.jpeg')) + 
                                 list(base_dir.glob('*.png')))
    
    if not available_covers:
        raise RuntimeError("No cover images found in clean_images/ or dataGen directory")
    
    logger.info("Found %d cover images", len(available_covers))
    
    # Setup output folder
    tests_folder = base_dir / "BOSS_sten_data"
    tests_folder.mkdir(parents=True, exist_ok=True)
    
    # augmented_folder = base_dir / "BOSS_sten_data_augmented"
    # augmented_folder.mkdir(parents=True, exist_ok=True)
    
    # clean_augmented_folder = base_dir / "BOSS_clean_data_augmented"
    # clean_augmented_folder.mkdir(parents=True, exist_ok=True)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "File Path"
    ws["B1"] = "Stegnography Applied?"
    ws["C1"] = "Payload Category"
    ws["D1"] = "Payload Size (bytes)"
    row_count = 2
    
    # Calculate payload distribution targets
    payload_targets = {}
    for category, config in PAYLOAD_DISTRIBUTION.items():
        count = int(TARGET_TOTAL * config['ratio'])
        payload_targets[category] = {
            'count': count,
            'secret_size': config['secret_size_bytes'],
            'description': config['description'],
            'generated': 0
        }
    
    logger.info("="*60)
    logger.info("PAYLOAD DISTRIBUTION TARGETS:")
    for category, target in payload_targets.items():
        logger.info("  %s: %d images (%s)", 
                   category.capitalize(), 
                   target['count'], 
                   target['description'])
    logger.info("  Total stego: %d", TARGET_TOTAL)
    logger.info("  Total clean: %d", len(available_covers))
    logger.info("="*60)
    
    # Shuffle covers and assess their capacities
    random.seed(42)
    shuffled_covers = list(available_covers)
    random.shuffle(shuffled_covers)
    
    # Estimate capacity for each cover and categorize
    logger.info("\nEstimating image capacities...")
    cover_capacities = []
    for cover in tqdm(shuffled_covers, desc="Analyzing covers"):
        capacity = estimate_image_capacity(cover)
        cover_capacities.append((cover, capacity))
    
    # Sort by capacity (descending)
    cover_capacities.sort(key=lambda x: x[1], reverse=True)
    
    # Distribute covers to payload categories based on their capacity
    cover_assignments = {'low': [], 'medium': [], 'high': []}
    
    # Assign covers starting with high payload (needs most capacity)
    for category in ['high', 'medium', 'low']:
        target = payload_targets[category]
        secret_size = target['secret_size']
        assigned = 0
        
        # Find covers with sufficient capacity
        for cover, capacity in cover_capacities:
            if capacity >= secret_size * 1.5:  # 50% safety margin
                cover_assignments[category].append(cover)
                assigned += 1
                if assigned >= target['count']:
                    break
        
        # Remove assigned covers from pool
        for cover in cover_assignments[category]:
            cover_capacities = [(c, cap) for c, cap in cover_capacities if c != cover]
        
        logger.info("Assigned %d covers to %s payload (needed %d, target size: %d bytes)", 
                   len(cover_assignments[category]), category, target['count'], secret_size)
    
    # Warn if we couldn't assign enough covers
    for category, covers in cover_assignments.items():
        if len(covers) < payload_targets[category]['count']:
            logger.warning("Only found %d suitable covers for %s payload (target: %d)", 
                          len(covers), category, payload_targets[category]['count'])
    
    # Initialize steghide tool
    steg = Steg_Gen()
    
    # Generate stego images by payload category
    logger.info("\nGenerating stego images with adaptive payload sizing...")
    
    for category, covers in cover_assignments.items():
        secret_size = payload_targets[category]['secret_size']
        logger.info("\n--- Generating %s payload stego images (%d bytes secret) ---", 
                   category, secret_size)
        
        category_pbar = tqdm(covers, desc=f"{category.capitalize()} payload")
        
        for cover_path in category_pbar:
            # Add original clean image
            ws.cell(row=row_count, column=1, value=str(cover_path.resolve()))
            ws.cell(row=row_count, column=2, value=False)
            ws.cell(row=row_count, column=3, value='N/A')
            ws.cell(row=row_count, column=4, value='N/A')
            row_count += 1
            
            # Generate stego image
            stego_filename = f"stego_{cover_path.stem}_{payload_targets[category]['generated']}.jpg"
            stego_path = tests_folder / stego_filename
            
            # Create secret data of specified size
            secret_data = generate_secret_data(secret_size, method='random')
            
            # Save secret to temp file
            with tempfile.NamedTemporaryFile(mode='wb', delete=False, suffix='.bin') as tmp:
                tmp.write(secret_data)
                secret_file = Path(tmp.name)
            
            try:
                # Embed with steghide
                steg.embed(
                    cover=str(cover_path),
                    secret=str(secret_file),
                    stego=str(stego_path),
                    password=PASSWORD
                )
                
                if stego_path.exists():
                    ws.cell(row=row_count, column=1, value=str(stego_path.resolve()))
                    ws.cell(row=row_count, column=2, value=True)
                    ws.cell(row=row_count, column=3, value=category)
                    ws.cell(row=row_count, column=4, value=secret_size)
                    row_count += 1
                    payload_targets[category]['generated'] += 1
                else:
                    logger.warning("Steghide failed for %s (output not created)", cover_path.name)
                    
            except RuntimeError as e:
                logger.warning("RuntimeError embedding %s (%d bytes secret): %s", 
                             cover_path.name, secret_size, str(e))
            except Exception as e:
                logger.error("Unexpected error embedding %s (%d bytes secret): %s", 
                           cover_path.name, secret_size, str(e))
            finally:
                # Clean up temp secret file
                try:
                    secret_file.unlink()
                except:
                    pass
    
    # Summary statistics
    logger.info("\n" + "="*60)
    logger.info("GENERATION SUMMARY")
    logger.info("="*60)
    logger.info("Total clean images: %d", len(available_covers))
    logger.info("Total stego images by category:")
    total_stego = 0
    for category, target in payload_targets.items():
        success_rate = 100 * target['generated'] / target['count'] if target['count'] > 0 else 0
        logger.info("  %s: %d / %d (%.1f%%)", 
                   category.capitalize(),
                   target['generated'],
                   target['count'],
                   success_rate)
        total_stego += target['generated']
    logger.info("Total stego images generated: %d", total_stego)
    logger.info("="*60)
    
    # Save workbook
    logger.info("\nSaving Excel manifest...")
    wb.save(EXCEL_OUT)
    
    # Calculate actual totals from workbook
    total_clean = 0
    total_stego_in_wb = 0
    category_counts = {'low': 0, 'medium': 0, 'high': 0}
    
    for row in ws.iter_rows(min_row=2):
        if row[1].value:  # Stego
            total_stego_in_wb += 1
            category = row[2].value
            if category in category_counts:
                category_counts[category] += 1
        else:
            total_clean += 1
    
    # Final summary
    logger.info("\n" + "="*60)
    logger.info("FINAL DATASET SUMMARY")
    logger.info("="*60)
    logger.info("Clean images: %d", total_clean)
    logger.info("Stego images by category:")
    for category in ['low', 'medium', 'high']:
        percentage = 100 * category_counts[category] / total_stego_in_wb if total_stego_in_wb > 0 else 0
        logger.info("  %s: %d (%.1f%%)", category.capitalize(), category_counts[category], percentage)
    logger.info("Total stego: %d", total_stego_in_wb)
    logger.info("Total images: %d", total_clean + total_stego_in_wb)
    logger.info("Stego/Clean ratio: %.2f", total_stego_in_wb / total_clean if total_clean > 0 else 0)
    logger.info("="*60)
    logger.info("Dataset manifest saved to: %s", EXCEL_OUT)
    logger.info("Generation complete!")


if __name__ == "__main__":
    main()    # Augment clean images

    # Augmented data to increase our custom dataset
    # total_clean_augmented = 0
    
    # if variants_per_clean > 0 and AUGMENT_CLEAN_RATIO > 0:
    #     logger.info("Augmenting clean images...")
        
    #     # Select which clean images to augment
    #     clean_to_augment = random.sample(available_covers, num_clean_to_augment)
        
    #     for cover_path in tqdm(clean_to_augment, desc='Augmenting clean', unit='image'):
    #         for aug_idx in range(variants_per_clean):
    #             aug_name = f"aug_clean_{aug_idx}_{cover_path.stem}{cover_path.suffix}"
    #             aug_path = clean_augmented_folder / aug_name
                
    #             try:
    #                 augmentor.augment_and_save(
    #                     cover_path, 
    #                     aug_path, 
    #                     transforms='random', 
    #                     mild=AUGMENT_MILD
    #                 )
                    
    #                 if aug_path.exists():
    #                     ws.cell(row=row_count, column=1, value=str(aug_path.resolve()))
    #                     ws.cell(row=row_count, column=2, value=False)
    #                     row_count += 1
    #                     total_clean_augmented += 1
    #             except Exception:
    #                 pass  # Skip failed augmentations
    
    # # Save workbook
    # excel_path = base_dir / "BOSS_stego_training.xlsx"
    # wb.save(excel_path)
    
    # # Calculate final statistics
    # total_clean = num_covers + total_clean_augmented
    # total_stego = total_stego_created + total_augmented_stego
    # total_images = total_clean + total_stego
    # clean_percent = (total_clean / total_images * 100) if total_images > 0 else 0
    # stego_percent = (total_stego / total_images * 100) if total_images > 0 else 0
    
    # logger.info("=" * 60)
    # logger.info("Generation complete!")
    # logger.info("  CLEAN IMAGES:")
    # logger.info("    Original: %d", num_covers)
    # logger.info("    Augmented: %d", total_clean_augmented)
    # logger.info("    Total: %d (%.1f%%)", total_clean, clean_percent)
    # logger.info("  STEGO IMAGES:")
    # logger.info("    Base: %d", total_stego_created)
    # logger.info("    Augmented: %d", total_augmented_stego)
    # logger.info("    Total: %d (%.1f%%)", total_stego, stego_percent)
    # logger.info("  TOTAL DATASET: %d images", total_images)
    # logger.info("  Excel saved to: %s", excel_path)
    # logger.info("=" * 60)
