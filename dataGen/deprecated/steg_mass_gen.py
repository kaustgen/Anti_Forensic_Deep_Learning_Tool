# Author: Kaleb Austgen
# Modify Date: 10-21-25
# Purpose:
"""   Goal of this script is to automate the creation of files modified by anti-forensic techniques
      Once the original file is generated the script will choose some anti-forensic techniques
      (Steganography, Encryption, Time-stomping, Data Wiping) and modify the file in a semi-randomized way
      It will then label the original file as original, and the modifed file with the specific technique used
      Once the data is labeled it will add it to a vector database
      The script will generate 100,000 different files, with half of them being modified through data-wiping
      and the other half being modified with steganography

      Generated Adversarial Networks, Convolutional Neural Networks are highly proficient in hiding images, 
      perhaps we can train a model that can counter these?

      Goal: Generate a series of files made by CNNs and other advanced systems and create a system that
      can detect it

      Must be able to detect -
            Least Significant Bit (LSB)
            

      https://sheriffjbabu.medium.com/python-ai-for-steganography-862e732cd3e0 - source
"""
from fpdf import FPDF
import random
from steg_gen import Steg_Gen
import logging
from pathlib import Path
from openpyxl import Workbook
from pathlib import Path
import shutil
from tqdm import tqdm

# pdf = FPDF()
# pdf.add_page()
# pdf.set_font("Arial", size=12)
# pdf.cell(200, 10, txt="Hello, this is a simple sentence in a PDF!", ln=True, align="L")

# pdf.output("hello.pdf")

def generate_pdf(text, output_file):
      output_file = Path(output_file)
      # Ensure parent directory exists before writing the PDF
      output_file.parent.mkdir(parents=True, exist_ok=True)

      pdf = FPDF()
      pdf.add_page()
      pdf.set_font("Arial", size=12)
      pdf.cell(200, 10, txt=text, ln=True, align="L")

      pdf.output(str(output_file))

steg = Steg_Gen()

# Password for steghide
PASSWORD = 'five'

# Generate 30 short sentences. Include a few with Username / Password / Email per request.
sentences = [
      "This is a short note.",
      "Please review the attachment.",
      "System check complete.",
      "User reported no issues.",
      "Temporary token issued.",
      "Meeting at noon.",
      "Backup completed successfully.",
      "Refer to the log file.",
      "Operation completed.",
      "Low disk space warning.",
      "New device connected.",
      "Configuration updated.",
      "Restart required.",
      "Security scan passed.",
      "Pending approval needed.",
      "Contact support if needed.",
      "Last login was yesterday.",
      "Session expired - please re-authenticate.",
      "Username: admin_user",            # contains 'Username'
      "Password: hunter2",               # contains 'Password'
      "Email: user@example.com",         # contains 'Email'
      "Temporary access granted.",
      "Configuration saved.",
      "Automatic update scheduled.",
      "License check OK.",
      "New message received.",
      "Sync completed.",
      "User profile updated.",
      "System maintenance planned.",
      "Reference ID: 12345",
      "Alert: unusual activity detected.",
      "Please change your password regularly.",
      "New user registered.",
      "Disk usage at 75%.",
      "Auto-save completed.",
      "Connection timed out.",
      "Service restarted successfully.",
      "Credential update required.",
      "User preferences synced.",
      "Temporary password: tmp1234",   # contains 'Password'
      "Contact: admin@example.com",    # contains 'Email'
      "Account Username: test_user",   # contains 'Username'
      "Build completed.",
      "Testing environment ready.",
      "Cache cleared.",
      "Dependency updated.",
      "Run ID: abcde-12345.",
      "Heartbeat OK.",
      "Feature flag toggled.",
      "Maintenance window scheduled."
]

# Discover cover images automatically. Prefer an 'Images' subfolder if present.
base_dir = Path(__file__).parent
# Look for both 'Images' and 'images' (user might have different capitalization)
images_dir = base_dir / 'clean_images'
images_dir_lower = base_dir / 'clean_images'
available_covers = []
for d in (images_dir, images_dir_lower):
      if d.exists() and d.is_dir():
            available_covers.extend(sorted(list(d.glob('*.jpg')) + list(d.glob('*.jpeg')) + list(d.glob('*.png'))))

# If none found in Images/, look in the dataGen folder itself
if not available_covers:
      available_covers = sorted(list(base_dir.glob('*.jpg')) + list(base_dir.glob('*.jpeg')) + list(base_dir.glob('*.png')))

# fallback to some expected filenames next to script
if not available_covers:
      available_covers = [base_dir / f for f in ["cover_falls.jpg", "cover_boat.jpg", "cover_girl.jpg", "cover_house.jpg"] if (base_dir / f).exists()]

if not available_covers:
      raise RuntimeError("No cover images found in Images/ or dataGen directory; please add jpg/png cover images.")

# Create workbook and headers
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

wb = Workbook()
ws = wb.active
ws["A1"] = "File Path"
ws["B1"] = "Stegnography Applied?"

# Start writing data on row 2 (row 1 is header)
row_count = 2
# Use script directory as the base so the script behaves the same regardless of CWD
base_dir = Path(__file__).parent
dataGen_folder = base_dir
tests_folder = dataGen_folder / "sten_data"
tests_folder.mkdir(parents=True, exist_ok=True)

# We'll generate N items per cover (total = N * num_covers)
# Compute number per cover to reach ~2000 total images
TARGET_TOTAL = 5000
num_covers = len(available_covers)
N_per_cover = (TARGET_TOTAL + num_covers - 1) // num_covers
logger.info("Found %d covers, generating %d items per cover (target ~%d)", num_covers, N_per_cover, TARGET_TOTAL)
total_counter = 0

for cover_path in tqdm(available_covers, desc='Generating data'):

      non_stego = cover_path

      # Add clean image path once (the non_stego file is guaranteed to exist)
      ws.cell(row=row_count, column=1, value=str(non_stego.resolve()))
      ws.cell(row=row_count, column=2, value=False)
      row_count += 1
      for i in range(N_per_cover):
            # Build a secret text by concatenating 4 random sentences and repeating that block 4 times
            block = ' '.join(random.choice(sentences) for _ in range(4))
            text = ' '.join([block] * 4)

            # sanitize a short name for filenames
            short_name = f"{cover_path.stem}_{i}"

            secret_pdf = tests_folder / f"og_{short_name}.pdf"
            stego_out = tests_folder / f"stego_{short_name}.jpg"
            extracted = tests_folder / f"rec_{short_name}.pdf"

            # generate PDF
            generate_pdf(text, secret_pdf)

            # call steg helper to create stego image
            try:
                  steg.create(str(non_stego), str(secret_pdf), str(stego_out), str(extracted), PASSWORD)
            except Exception as e:
                  # If steg.create expects Path objects instead of strings, try that
                  try:
                        steg.create(non_stego, secret_pdf, stego_out, extracted, PASSWORD)
                  except Exception as e2:
                        # record failure and skip
                        logger = logging.getLogger(__name__)
                        logger.error("steg.create failed for %s -> %s: %s; %s", non_stego, stego_out, e, e2)
                        continue

            # Add stego image per N
            ws.cell(row=row_count, column=1, value=str(stego_out.resolve()))
            ws.cell(row=row_count, column=2, value=True)
            row_count += 1

            total_counter += 1
            if total_counter % 100 == 0:
                  logger.info("Generated %d stego pairs...", total_counter)

wb.save(dataGen_folder / "stego_training.xlsx")
logger.info("Found %d covers, generating %d items per cover (target ~%d)", num_covers, N_per_cover, TARGET_TOTAL)
logger.info("Wrote excel to %s", str(dataGen_folder / "stego_training.xlsx"))


#   cover: File path of item that the secret will be hidden in
#   secret: File path of the secret file insert
#   stego: Output cover that contains the secret
#   extracted: Secret from setgo
#   password: Encryption password


