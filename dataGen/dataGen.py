# Author: Kaleb Austgen
# Modify Date: 9-25-25
# Purpose:
"""   Goal of this script is to automate the creation of files modified by anti-forensic techniques
      Once the original file is generated the script will choose some anti-forensic techniques
      (Steganography, Encryption, Time-stomping, Data Wiping) and modify the file in a semi-randomized way
      It will then label the original file as original, and the modifed file with the specific technique used
      Once the data is labeled it will add it to a vector database
      The script will generate 100,000 different files, with half of them being modified through data-wiping
      and the other half being modified with steganography

      Generated Adversarial Networks, Convolutional Neural Networks are highly proficient in hiding images, 
      perhaps we can train a model that can counter these?

      Goal: Generate a series of files made by CNNs and other advanced systems and create a system that
      can detect it

      Must be able to detect -
            Least Significant Bit (LSB)
            

      https://sheriffjbabu.medium.com/python-ai-for-steganography-862e732cd3e0 - source
"""
from fpdf import FPDF
import random
from steg_gen import Steg_Gen
from openpyxl import Workbook
from pathlib import Path

# pdf = FPDF()
# pdf.add_page()
# pdf.set_font("Arial", size=12)
# pdf.cell(200, 10, txt="Hello, this is a simple sentence in a PDF!", ln=True, align="L")

# pdf.output("hello.pdf")

def generate_pdf(text, output_file):
      output_file = Path(output_file)
      # Ensure parent directory exists before writing the PDF
      output_file.parent.mkdir(parents=True, exist_ok=True)

      pdf = FPDF()
      pdf.add_page()
      pdf.set_font("Arial", size=12)
      pdf.cell(200, 10, txt=text, ln=True, align="L")

      pdf.output(str(output_file))

sentence = ['My secret', 'My super duper secret', 'joly beans, holy greens, and billy jeen', 'woah oh ohhhhhh']

cover = ['cover_falls', 'cover_boat', 'cover_girl', 'cover_house']

steg = Steg_Gen()

PASSWORD = 'five'

# Create a new excel workbook

wb = Workbook()
ws = wb.active
ws["A1"] = "File Path"
ws["B1"] = "Stegnography Applied?"

row_count = 1
# Generate 100 different items

# Use script directory as the base so the script behaves the same regardless of CWD
base_dir = Path(__file__).parent
dataGen_folder = base_dir
tests_folder = dataGen_folder / "gen_data"
for i in range(0, 50):

      # Randomly pick a secret and an image cover
      cover_pick = cover[random.randint(0,3)]

      sentence_pick = sentence[random.randint(0, 3)]
      sentence_pick = sentence_pick.replace(" ", "_")

      #secret_pdf = f"dataGen/tests/og_{sentence_pick}_{i}.pdf"
      secret_pdf = dataGen_folder / tests_folder / f"og_{sentence_pick}_{i}.pdf"

      #stego = f"dataGen/tests/{sentence_pick}_{cover_pick}_{i}.jpg"
      stego = dataGen_folder / tests_folder / f"{sentence_pick}_{cover_pick}_{i}.jpg"

      #extracted = f"dataGen/tests/rec_{sentence_pick}_{i}.pdf"
      extracted = dataGen_folder / tests_folder / f"rec_{sentence_pick}_{i}.pdf"

      generate_pdf(sentence_pick, secret_pdf)

      #non_stego = f"dataGen/{cover_pick}.jpg"
      non_stego = dataGen_folder / f"{cover_pick}.jpg"

      steg.create(non_stego, secret_pdf, stego, extracted, PASSWORD)

      ws.cell(row=row_count, column=1, value=str(non_stego))
      ws.cell(row=row_count, column=2, value=False)
      row_count += 1
      ws.cell(row=row_count, column=1, value=str(stego))
      ws.cell(row=row_count, column=2, value=True)
      row_count += 1

wb.save(dataGen_folder / "stego_training.xlsx")

#   cover: File path of item that the secret will be hidden in
#   secret: File path of the secret file insert
#   stego: Output cover that contains the secret
#   extracted: Secret from setgo
#   password: Encryption password


