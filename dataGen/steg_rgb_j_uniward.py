#!/usr/bin/env python3
"""
Generate stego images from RGB augmented covers using J-UNIWARD.

Embeds steganographic data into augmented RGB images at mixed rates
(0.1-0.4 bpnzAC) to match the BOSS dataset distribution. 

IMPORTANT: RGB images are kept in full color to provide real-world
diverse training data. This is the entire purpose of this dataset -
to expose the model to color images with diverse content, not just
grayscale nature scenes from BOSS.

J-UNIWARD operates on JPEG DCT coefficients channel-by-channel, so
it works perfectly fine with RGB images without conversion.

Author: Kaleb Austgen
Date: October 30, 2025
"""

import logging
from pathlib import Path
from openpyxl import Workbook
from tqdm import tqdm
import random
from j_uniward import JUNIWARDEmbedder
from PIL import Image

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ============================================================
# CONFIGURATION
# ============================================================

BASE_DIR = Path(__file__).parent

# Input: Augmented RGB covers
COVER_DIR = BASE_DIR / 'clean_images'

# Output: Stego images and metadata
STEGO_DIR = BASE_DIR / 'rgb_stego'
EXCEL_OUT = BASE_DIR / 'RGB_stego_metadata.xlsx'

# Match BOSS embedding rate distribution
# This ensures RGB and BOSS datasets have same stego characteristics
EMBEDDING_RATES = {
    'low': {
        'ratio': 0.25,
        'bpnzAC': 0.1,
        'description': 'Low embedding rate (0.1 bpnzAC)'
    },
    'medium': {
        'ratio': 0.25,
        'bpnzAC': 0.2,
        'description': 'Medium embedding rate (0.2 bpnzAC)'
    },
    'high': {
        'ratio': 0.25,
        'bpnzAC': 0.3,
        'description': 'High embedding rate (0.3 bpnzAC)'
    },
    'standard': {
        'ratio': 0.25,
        'bpnzAC': 0.4,
        'description': 'Standard embedding rate (0.4 bpnzAC)'
    }
}

# Random seed for reproducibility
RANDOM_SEED = 42

# ============================================================
# HELPER FUNCTIONS
# ============================================================

def ensure_rgb_format(img_path):
    """
    Ensure image is in RGB format and 512×512 size for J-UNIWARD embedding.
    
    Converts RGBA to RGB if needed (removes alpha channel).
    Keeps RGB images as-is to preserve color information.
    Resizes all images to 512×512 to match BOSS dataset dimensions.
    
    Args:
        img_path: Path to input image
        
    Returns:
        PIL Image in RGB format, resized to 512×512
    """
    img = Image.open(img_path)
    
    # Convert RGBA to RGB (remove alpha channel)
    if img.mode == 'RGBA':
        rgb_img = Image.new('RGB', img.size, (255, 255, 255))
        rgb_img.paste(img, mask=img.split()[3])  # Use alpha as mask
        img = rgb_img
    
    # Convert grayscale to RGB (rare, but handle it)
    elif img.mode in ['L', '1']:
        img = img.convert('RGB')
    
    # Convert other modes to RGB
    elif img.mode != 'RGB':
        img = img.convert('RGB')
    
    # Resize to 512×512 to match BOSS dimensions
    if img.size != (512, 512):
        img = img.resize((512, 512), Image.LANCZOS)
    
    return img


# ============================================================
# MAIN EMBEDDING PIPELINE
# ============================================================

def main():
    """Generate stego images from augmented RGB covers."""
    
    logger.info("="*70)
    logger.info("RGB J-UNIWARD STEGANOGRAPHY EMBEDDING")
    logger.info("="*70)
    
    # Find all cover images
    cover_images = []
    for ext in ['*.jpg', '*.jpeg', '*.png', '*.webp']:
        cover_images.extend(COVER_DIR.glob(ext))
    
    cover_images = sorted(cover_images)
    
    if not cover_images:
        logger.error(f"No cover images found in {COVER_DIR}")
        logger.error("Run augment_rgb_covers.py first!")
        return
    
    logger.info(f"Cover directory:      {COVER_DIR}")
    logger.info(f"Stego directory:      {STEGO_DIR}")
    logger.info(f"Excel output:         {EXCEL_OUT}")
    logger.info(f"Found covers:         {len(cover_images)}")
    logger.info(f"Random seed:          {RANDOM_SEED}")
    logger.info("")
    logger.info("Embedding rate distribution (matching BOSS):")
    for category, config in EMBEDDING_RATES.items():
        count = int(len(cover_images) * config['ratio'])
        logger.info(f"  {category.capitalize()}: {count} images @ {config['bpnzAC']} bpnzAC")
    logger.info("="*70)
    
    # Create output directories
    STEGO_DIR.mkdir(parents=True, exist_ok=True)
    
    # Shuffle and assign embedding rates
    random.seed(RANDOM_SEED)
    random.shuffle(cover_images)
    
    rate_assignments = {}
    remaining_images = cover_images.copy()
    
    for category, config in EMBEDDING_RATES.items():
        count = int(len(cover_images) * config['ratio'])
        rate_assignments[category] = remaining_images[:count]
        remaining_images = remaining_images[count:]
    
    # Create Excel workbook with BOSS-compatible structure
    wb = Workbook()
    ws = wb.active
    
    # Match BOSS Excel column structure exactly
    ws["A1"] = "File Path"
    ws["B1"] = "Stegnography Applied?"
    ws["C1"] = "Payload Category"
    ws["D1"] = "Payload Size (bytes)"
    ws["E1"] = "Payload (bpp AC DCT)"
    ws["F1"] = "Payload (bytes)"
    ws["G1"] = "Payload (bits)"
    ws["H1"] = "Image Dimensions"
    ws["I1"] = "Non-zero AC DCT"
    ws["J1"] = "RGB"  # New column to identify RGB vs grayscale
    
    row_count = 2
    
    # Statistics
    stats = {
        'total_covers': 0,
        'total_stego': 0,
        'failed': 0,
        'by_rate': {cat: 0 for cat in EMBEDDING_RATES.keys()}
    }
    
    # Process each embedding rate
    logger.info("\nStarting embedding...")
    for category, covers in rate_assignments.items():
        bpnzAC = EMBEDDING_RATES[category]['bpnzAC']
        
        logger.info(f"\nProcessing {category} rate ({bpnzAC} bpnzAC): {len(covers)} images")
        
        embedder = JUNIWARDEmbedder(alpha=bpnzAC)
        
        for cover_path in tqdm(covers, desc=f"{category.capitalize()} rate", unit="img"):
            try:
                # Ensure image is in RGB format and 512×512 (resize/convert if needed)
                rgb_img = ensure_rgb_format(cover_path)
                
                # Overwrite original image in clean_images if it was modified
                # This ensures all images in clean_images are 512×512 RGB JPEGs
                rgb_img.save(cover_path, 'JPEG', quality=95)
                rgb_img.close()
                
                # Add clean cover image entry (matches BOSS format)
                ws.cell(row=row_count, column=1, value=str(cover_path.resolve()))
                ws.cell(row=row_count, column=2, value=False)  # Stegnography Applied?
                ws.cell(row=row_count, column=3, value='N/A')  # Payload Category
                ws.cell(row=row_count, column=4, value='N/A')  # Payload Size (bytes)
                ws.cell(row=row_count, column=5, value='N/A')  # Payload (bpp AC DCT)
                ws.cell(row=row_count, column=6, value='N/A')  # Payload (bytes)
                ws.cell(row=row_count, column=7, value='N/A')  # Payload (bits)
                ws.cell(row=row_count, column=8, value="512×512")  # Image Dimensions (now standardized)
                ws.cell(row=row_count, column=9, value='N/A')  # Non-zero AC DCT
                ws.cell(row=row_count, column=10, value=True)  # RGB = True
                row_count += 1
                stats['total_covers'] += 1
                
                # Use the (now standardized) cover image for embedding
                cover_for_embedding = cover_path
                
                # Generate stego filename
                stego_filename = f"{cover_path.stem}_stego_{category}.jpg"
                stego_path = STEGO_DIR / stego_filename
                
                # Embed steganography (RGB image stays RGB!)
                embed_stats = embedder.embed(
                    cover_path=str(cover_for_embedding),
                    stego_path=str(stego_path),
                    payload_bytes=None  # Auto-calculate from bpnzAC
                )
                
                if stego_path.exists():
                    # Add stego image entry (matches BOSS format)
                    ws.cell(row=row_count, column=1, value=str(stego_path.resolve()))
                    ws.cell(row=row_count, column=2, value=True)  # Stegnography Applied?
                    ws.cell(row=row_count, column=3, value=category)  # Payload Category
                    ws.cell(row=row_count, column=4, value=embed_stats.get('payload_bytes', 0))  # Payload Size (bytes)
                    ws.cell(row=row_count, column=5, value=f"{embed_stats.get('bpnzAC', bpnzAC):.6f}")  # Payload (bpp AC DCT)
                    ws.cell(row=row_count, column=6, value=embed_stats.get('payload_bytes', 0))  # Payload (bytes)
                    ws.cell(row=row_count, column=7, value=embed_stats.get('payload_bits', 0))  # Payload (bits)
                    ws.cell(row=row_count, column=8, value=f"{embed_stats.get('width', 0)}x{embed_stats.get('height', 0)}")  # Image Dimensions
                    ws.cell(row=row_count, column=9, value=embed_stats.get('nzAC', 0))  # Non-zero AC DCT
                    ws.cell(row=row_count, column=10, value=True)  # RGB = True
                    row_count += 1
                    
                    stats['total_stego'] += 1
                    stats['by_rate'][category] += 1
                else:
                    logger.warning(f"Stego file not created: {stego_filename}")
                    stats['failed'] += 1
                    
            except Exception as e:
                logger.error(f"Failed to process {cover_path.name}: {e}")
                stats['failed'] += 1
    
    # Save Excel file
    try:
        wb.save(EXCEL_OUT)
        logger.info(f"\n✓ Excel metadata saved: {EXCEL_OUT}")
    except Exception as e:
        logger.error(f"Failed to save Excel file: {e}")
    
    # Summary
    logger.info("\n" + "="*70)
    logger.info("EMBEDDING COMPLETE")
    logger.info("="*70)
    logger.info(f"Cover images:         {stats['total_covers']}")
    logger.info(f"Stego images:         {stats['total_stego']}")
    logger.info(f"Total dataset:        {stats['total_covers'] + stats['total_stego']} (all RGB)")
    logger.info(f"Failed:               {stats['failed']}")
    logger.info("")
    logger.info("Stego images by rate:")
    for category in EMBEDDING_RATES.keys():
        logger.info(f"  {category.capitalize()}: {stats['by_rate'][category]}")
    logger.info("")
    logger.info(f"Stego directory:      {STEGO_DIR}")
    logger.info(f"Excel metadata:       {EXCEL_OUT}")
    logger.info("="*70)
    logger.info("\nNext step: Run combine_boss_rgb.py to merge with BOSS dataset")
    logger.info("="*70)


if __name__ == '__main__':
    main()
