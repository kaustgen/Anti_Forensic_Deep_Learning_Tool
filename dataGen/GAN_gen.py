# Code generated for research from this article
# @article{zhang2019steganogan,
#   title={SteganoGAN: High Capacity Image Steganography with GANs},
#   author={Zhang, Kevin Alex and Cuesta-Infante, Alfredo and Veeramachaneni, Kalyan},
#   journal={arXiv preprint arXiv:1901.03892},
#   year={2019},
#   url={https://arxiv.org/abs/1901.03892}
# }

#!/usr/bin/env python3
"""
Generate GAN-based steganography using SteganoGAN for testing.

Creates adversarial stego images that are harder to detect than J-UNIWARD,
formatted identically to BOSS dataset for compatibility with existing pipeline.

Author: Kaleb Austgen
Date: November 1, 2025
"""

import logging
from pathlib import Path
from openpyxl import Workbook
from tqdm import tqdm
import random
import numpy as np
from PIL import Image
import torch
import string
import time
import secrets

# Monkey-patch torch.optim.Adam to handle legacy state loading
import torch.optim as optim

_original_adam_setstate = optim.Adam.__setstate__

def _patched_adam_setstate(self, state):
    """Wrapper for Adam.__setstate__ that handles missing defaults attribute."""
    # Initialize defaults if not present (for legacy checkpoints)
    if not hasattr(self, 'defaults'):
        self.defaults = {}
    if not hasattr(self, 'state'):
        self.state = {}
    if not hasattr(self, 'param_groups'):
        self.param_groups = []
    
    try:
        _original_adam_setstate(self, state)
    except AttributeError:
        # Fallback: manually set state for legacy models
        self.__dict__.update(state)
        if 'defaults' not in self.__dict__:
            self.defaults = {}

optim.Adam.__setstate__ = _patched_adam_setstate

# Monkey-patch torch.load to handle SteganoGAN models with PyTorch 2.6+
_original_torch_load = torch.load

def _patched_torch_load(f, *args, **kwargs):
    """Wrapper for torch.load that sets weights_only=False for SteganoGAN models."""
    # If weights_only is not explicitly set, set it to False for backward compatibility
    if 'weights_only' not in kwargs:
        kwargs['weights_only'] = False
    return _original_torch_load(f, *args, **kwargs)

torch.load = _patched_torch_load

# Import SteganoGAN (after patching)
from steganogan import SteganoGAN

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ============================================================
# CONFIGURATION
# ============================================================

BASE_DIR = Path(__file__).parent
EXCEL_OUT = BASE_DIR / 'BOSS_steganogan_metadata.xlsx'
TARGET_STEGO_IMAGES = 100  # Subset for testing


def calculate_equivalent_bpp(bpnzac_target, image_size=(512, 512)):
    """
    Calculate SteganoGAN bpp equivalent to J-UNIWARD bpnzAC.
    
    J-UNIWARD embeds in non-zero AC DCT coefficients (only ~3-10% of total coefficients).
    SteganoGAN embeds across all pixels, so we need much lower bpp for equivalent capacity.
    
    Args:
        bpnzac_target: Target J-UNIWARD rate (e.g., 0.1, 0.2, 0.3, 0.4)
        image_size: (width, height) tuple
    
    Returns:
        float: Equivalent bits per pixel for SteganoGAN
    """
    width, height = image_size
    total_pixels = width * height
    
    # Estimate non-zero AC DCT coefficients
    # For JPEG with 8×8 blocks:
    num_blocks = (width // 8) * (height // 8)
    ac_coeffs_per_block = 63  # Exclude DC coefficient
    
    # Conservative estimate: ~10% of AC coefficients are non-zero in typical JPEG
    # (This varies by image complexity and JPEG quality, 5-15% is typical)
    nonzero_ac_ratio = 0.10
    estimated_nonzero_ac = num_blocks * ac_coeffs_per_block * nonzero_ac_ratio
    
    # Calculate payload in bits for J-UNIWARD at target rate
    payload_bits = estimated_nonzero_ac * bpnzac_target
    
    # Convert to bpp for SteganoGAN (spread across ALL pixels)
    equivalent_bpp = payload_bits / total_pixels
    
    logger.info(f"  J-UNIWARD {bpnzac_target:.1f} bpnzAC ≈ SteganoGAN {equivalent_bpp:.4f} bpp")
    logger.info(f"  Estimated payload: {payload_bits:.0f} bits ({payload_bits/8:.0f} bytes)")
    
    return equivalent_bpp


# SteganoGAN Dense model at J-UNIWARD EQUIVALENT embedding rates
# NOTE: These bpp values are much smaller than typical SteganoGAN usage (0.4-4.0 bpp)
# because we're matching J-UNIWARD's capacity, not maximizing SteganoGAN capacity
STEGANOGAN_CONFIGS = {
    'low': {
        'ratio': 0.25,
        'model': 'dense',
        'bits_per_pixel': calculate_equivalent_bpp(0.1),  # ≈0.004 bpp
        'bpnzac_equivalent': 0.1,
        'description': 'SteganoGAN Dense (Low - equivalent to 0.1 bpnzAC)'
    },
    'medium': {
        'ratio': 0.25,
        'model': 'dense',
        'bits_per_pixel': calculate_equivalent_bpp(0.2),  # ≈0.008 bpp
        'bpnzac_equivalent': 0.2,
        'description': 'SteganoGAN Dense (Medium - equivalent to 0.2 bpnzAC)'
    },
    'high': {
        'ratio': 0.25,
        'model': 'dense',
        'bits_per_pixel': calculate_equivalent_bpp(0.3),  # ≈0.012 bpp
        'bpnzac_equivalent': 0.3,
        'description': 'SteganoGAN Dense (High - equivalent to 0.3 bpnzAC)'
    },
    'standard': {
        'ratio': 0.25,
        'model': 'dense',
        'bits_per_pixel': calculate_equivalent_bpp(0.4),  # ≈0.015 bpp
        'bpnzac_equivalent': 0.4,
        'description': 'SteganoGAN Dense (Standard - equivalent to 0.4 bpnzAC)'
    }
}


class SteganoGANEmbedder:
    """Wrapper for SteganoGAN embedding with metadata tracking."""
    
    def __init__(self, model_name='basic'):
        """
        Initialize SteganoGAN model.
        
        Args:
            model_name: 'basic', 'dense', or 'residual'
        """
        logger.info(f"Loading SteganoGAN model: {model_name}")
        self.model = SteganoGAN.load(architecture=model_name)
        self.device = 'cuda' if torch.cuda.is_available() else 'cpu'
        logger.info(f"Using device: {self.device}")
    
    def embed(self, cover_path, stego_path, bits_per_pixel=2.0):
        """
        Embed random message using SteganoGAN.
        
        Args:
            cover_path: Path to cover image
            stego_path: Path to save stego image
            bits_per_pixel: Embedding rate (higher = more capacity, harder to detect)
        
        Returns:
            dict: Statistics about embedding
        """
        # Get image dimensions for stats
        cover_path_obj = Path(cover_path)
        cover = Image.open(cover_path)
        width, height = cover.size
        
        # Resize large images to 512×512 to avoid CUDA OOM
        max_size = 512
        temp_cover_path = None
        if width > max_size or height > max_size:
            logger.warning(f"Resizing {cover_path_obj.name} from {width}×{height} to {max_size}×{max_size}")
            cover = cover.resize((max_size, max_size), Image.LANCZOS)
            width, height = max_size, max_size
            # Save resized version temporarily
            temp_cover_path = cover_path_obj.parent / f"temp_{cover_path_obj.name}"
            cover.save(temp_cover_path, quality=95)
            cover_path_to_use = str(temp_cover_path)
        else:
            cover_path_to_use = str(cover_path)
        
        cover.close()
        
        # Calculate message length
        total_pixels = width * height
        message_bits = int(total_pixels * bits_per_pixel)
        message_bytes = message_bits // 8

        # Generate truly random seed using secrets module (cryptographically secure)
        seed = secrets.randbits(64)  # 64-bit random number
        random.seed(seed)
        
        # Generate random text message (SteganoGAN expects text, not binary)
        # Use printable ASCII characters for realistic text
        message_chars = message_bytes  # Approximate: 1 char ≈ 1 byte
        message = ''.join(random.choices(string.ascii_letters + string.digits + ' ', k=message_chars))
        
        # Reset random state to avoid affecting other random calls
        random.seed()
        
        # Embed using SteganoGAN
        try:
            # SteganoGAN.encode() signature: (cover_path, output_path, text)
            # It expects path strings, not PIL Image objects
            self.model.encode(cover_path_to_use, str(stego_path), message)
            
            # Clean up temp file if created
            if temp_cover_path and temp_cover_path.exists():
                temp_cover_path.unlink(missing_ok=True)
            
            # Verify output was created
            if not Path(stego_path).exists():
                raise RuntimeError(f"SteganoGAN failed to create output: {stego_path}")
            
            # Convert RGB output to grayscale to match J-UNIWARD training data
            stego_rgb = Image.open(stego_path)
            stego_gray = stego_rgb.convert('L')  # Convert to grayscale
            stego_gray.save(stego_path, quality=95)  # Overwrite with grayscale version
            stego_rgb.close()
            stego_gray.close()
            
            # Clear CUDA cache to prevent memory buildup
            if torch.cuda.is_available():
                torch.cuda.empty_cache()
            
            # Calculate statistics (approximate for GAN)
            stats = {
                'payload_bits': message_bits,
                'payload_bytes': message_bytes,
                'bpp': bits_per_pixel,
                'dimensions': f"{width}×{height}",
                'total_pixels': total_pixels,
                'method': 'SteganoGAN',
                'model': self.model.__class__.__name__
            }
            
            return stats
            
        except Exception as e:
            # Clean up temp file on error
            if temp_cover_path and temp_cover_path.exists():
                temp_cover_path.unlink(missing_ok=True)
            
            # Clear CUDA cache on error
            if torch.cuda.is_available():
                torch.cuda.empty_cache()
            
            logger.error(f"SteganoGAN embedding failed: {e}")
            raise


def main():
    """Generate GAN-based stego dataset."""
    
    # Find cover images from BOSS dataset
    # Using BOSS ensures clean images have same characteristics as training data
    images_dir = BASE_DIR / 'BOSS1.01_Dataset'
    available_covers = []
    
    if images_dir.exists() and images_dir.is_dir():
        # Get all BOSS JPEG files and take a subset to avoid training overlap
        all_images = sorted(list(images_dir.glob('*.jpg')) + 
                           list(images_dir.glob('*.jpeg')))
        # Use images 10000-10099 (avoid the first 10000 used in training)
        available_covers = all_images[10000:10100] if len(all_images) > 10100 else all_images[:100]
    
    if not available_covers:
        raise RuntimeError(f"No cover images found in {images_dir}")
    
    logger.info("="*70)
    logger.info("STEGANOGAN GAN-BASED STEGO GENERATION")
    logger.info("="*70)
    logger.info("Found %d cover images (from BOSS dataset, images 10000-10099)", len(available_covers))
    logger.info("Target stego images: %d", TARGET_STEGO_IMAGES)
    logger.info("")
    logger.info("PAYLOAD CALCULATION (matching J-UNIWARD capacity):")
    logger.info("  J-UNIWARD embeds in ~10%% of DCT coefficients (non-zero AC)")
    logger.info("  SteganoGAN embeds across ALL pixels")
    logger.info("  Therefore: bpp values are much lower than typical SteganoGAN usage")
    logger.info("")
    logger.info("Model: Dense (DenseNet-based, highest quality)")
    logger.info("="*70)
    
    # Setup output folder
    stego_folder = BASE_DIR / "BOSS_steganogan_data"
    stego_folder.mkdir(parents=True, exist_ok=True)
    
    # Create workbook (same format as steg_boss_j_uniward.py)
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "File Path"
    ws["B1"] = "Stegnography Applied?"
    ws["C1"] = "Payload Category"
    ws["D1"] = "Payload Size (bytes)"
    ws["E1"] = "Payload (bpp AC DCT)"
    ws["F1"] = "Payload (bytes)"
    ws["G1"] = "Payload (bits)"
    ws["H1"] = "Image Dimensions"
    ws["I1"] = "Non-zero AC DCT"
    ws["J1"] = "RGB"
    row_count = 2
    
    # Calculate targets for each config
    config_targets = {}
    for category, config in STEGANOGAN_CONFIGS.items():
        count = int(TARGET_STEGO_IMAGES * config['ratio'])
        config_targets[category] = {
            'count': count,
            'model': config['model'],
            'bpp': config['bits_per_pixel'],
            'bpnzac_equivalent': config['bpnzac_equivalent'],  # Add this field!
            'description': config['description'],
            'generated': 0
        }
    
    logger.info("\nSteganoGAN Configuration:")
    for category, target in config_targets.items():
        logger.info("  %s: %d images (%s)", 
                   category.replace('_', ' ').title(), 
                   target['count'], 
                   target['description'])
    logger.info("="*70)
    
    # Shuffle covers once
    random.seed(42)  # Set seed for reproducibility
    shuffled_covers = list(available_covers)
    random.shuffle(shuffled_covers)
    random.seed()  # Reset seed for random message generation
    
    # STEP 2 FIX: Use UNIQUE covers for each payload rate (no reuse)
    # This ensures each rate gets different images with different stego patterns
    cover_assignments = {cat: [] for cat in STEGANOGAN_CONFIGS.keys()}
    cover_idx = 0
    
    for category in STEGANOGAN_CONFIGS.keys():
        target_count = config_targets[category]['count']
        # Get UNIQUE covers for this category (no modulo reuse!)
        if cover_idx + target_count > len(shuffled_covers):
            logger.warning(f"Not enough unique covers! Needed {cover_idx + target_count}, have {len(shuffled_covers)}")
            logger.warning("Some covers will be reused across categories.")
        
        for _ in range(target_count):
            # Use modulo only if we run out of covers
            cover_assignments[category].append(shuffled_covers[cover_idx % len(shuffled_covers)])
            cover_idx += 1
    
    logger.info(f"\nCover distribution: Using {cover_idx} cover assignments from {len(shuffled_covers)} unique images")
    if cover_idx > len(shuffled_covers):
        logger.info(f"  Note: {cover_idx - len(shuffled_covers)} covers will be reused")
    
    # Generate stego images by config
    logger.info("\nGenerating GAN-based stego images...")
    
    stats_summary = {
        'total_payload_bytes': 0,
        'total_payload_bits': 0
    }
    
    # Track loaded models to avoid reloading
    loaded_models = {}
    
    for category, covers in cover_assignments.items():
        model_name = config_targets[category]['model']
        bpp = config_targets[category]['bpp']
        
        logger.info("\n--- Generating %s ---", config_targets[category]['description'])
        
        # Load model once per architecture
        if model_name not in loaded_models:
            loaded_models[model_name] = SteganoGANEmbedder(model_name=model_name)
        
        embedder = loaded_models[model_name]
        category_pbar = tqdm(covers, desc=f"{category.replace('_', ' ').title()}")
        
        for cover_path in category_pbar:
            # Use ORIGINAL BOSS images as clean (no recompression, no modifications)
            # This ensures clean images are identical to training data
            
            # Add original BOSS image to Excel (no copying, no reprocessing)
            ws.cell(row=row_count, column=1, value=str(cover_path.resolve()))
            ws.cell(row=row_count, column=2, value=False)  # Stegnography Applied? = False
            ws.cell(row=row_count, column=3, value='N/A')
            ws.cell(row=row_count, column=4, value='N/A')  # Payload Size (bytes) - N/A for clean
            ws.cell(row=row_count, column=5, value='N/A')  # Payload (bpp AC DCT) - N/A for clean
            ws.cell(row=row_count, column=6, value='N/A')  # Payload (bytes) - N/A for clean
            ws.cell(row=row_count, column=7, value='N/A')  # Payload (bits) - N/A for clean
            ws.cell(row=row_count, column=8, value="512×512")  # BOSS standard size
            ws.cell(row=row_count, column=9, value='N/A')  # Non-zero AC DCT - N/A for clean
            ws.cell(row=row_count, column=10, value=False)  # RGB = False (grayscale)
            row_count += 1
            
            # Generate stego image
            stego_filename = f"steganogan_{category}_{config_targets[category]['generated']:04d}.jpg"
            stego_path = stego_folder / stego_filename
            
            try:
                # Embed with SteganoGAN
                stats = embedder.embed(
                    cover_path=str(cover_path),
                    stego_path=str(stego_path),
                    bits_per_pixel=bpp
                )
                
                if stego_path.exists():
                    ws.cell(row=row_count, column=1, value=str(stego_path.resolve()))
                    ws.cell(row=row_count, column=2, value=True)  # Stegnography Applied? = True
                    ws.cell(row=row_count, column=3, value=category)
                    ws.cell(row=row_count, column=4, value=stats['payload_bytes'])  # int, not string
                    # Store the equivalent bpnzAC for comparison with J-UNIWARD (as float, not string)
                    bpnzac_equiv = config_targets[category].get('bpnzac_equivalent', stats['bpp'])
                    ws.cell(row=row_count, column=5, value=bpnzac_equiv)  # float, not formatted string
                    ws.cell(row=row_count, column=6, value=stats['payload_bytes'])  # int, not string
                    ws.cell(row=row_count, column=7, value=stats['payload_bits'])  # int, not string
                    ws.cell(row=row_count, column=8, value=stats['dimensions'])  # string OK (e.g., "512×512")
                    ws.cell(row=row_count, column=9, value='N/A')  # Non-zero AC DCT (not applicable for GAN)
                    ws.cell(row=row_count, column=10, value=False)  # RGB = False (converted to grayscale)
                    row_count += 1
                    
                    config_targets[category]['generated'] += 1
                    stats_summary['total_payload_bytes'] += stats['payload_bytes']
                    stats_summary['total_payload_bits'] += stats['payload_bits']
                else:
                    logger.warning("SteganoGAN failed for %s (output not created)", Path(cover_path).name)
                    
            except Exception as e:
                logger.error("Error embedding %s: %s", Path(cover_path).name, str(e))
                continue
    
    # Summary statistics
    logger.info("\n" + "="*70)
    logger.info("GENERATION SUMMARY")
    logger.info("="*70)
    logger.info("Total clean images: %d", len(available_covers))
    logger.info("Total stego images by configuration:")
    total_stego = 0
    for category, target in config_targets.items():
        success_rate = 100 * target['generated'] / target['count'] if target['count'] > 0 else 0
        logger.info("  %s: %d / %d (%.1f%%)", 
                   category.replace('_', ' ').title(),
                   target['generated'],
                   target['count'],
                   success_rate)
        total_stego += target['generated']
    
    logger.info("Total stego images generated: %d", total_stego)
    
    if total_stego > 0:
        avg_payload_bytes = stats_summary['total_payload_bytes'] / total_stego
        avg_payload_bits = stats_summary['total_payload_bits'] / total_stego
        
        logger.info("\nEmbedding Statistics:")
        logger.info("  Average payload: %.1f bytes (%.1f bits)", avg_payload_bytes, avg_payload_bits)
    
    logger.info("="*70)
    
    # Save workbook
    logger.info("\nSaving Excel manifest...")
    wb.save(EXCEL_OUT)
    
    # Final summary
    logger.info("\n" + "="*70)
    logger.info("FINAL GAN STEGO DATASET SUMMARY")
    logger.info("="*70)
    logger.info("Dataset manifest: %s", EXCEL_OUT)
    logger.info("Stego images folder: %s", stego_folder)
    logger.info("Total images: %d", row_count - 2)
    logger.info("Format: Compatible with existing pipeline")
    logger.info("RGB: False (converted to grayscale to match J-UNIWARD training data)")
    logger.info("Source: BOSS dataset images 10000-10099 (separate from training set 0-9999)")
    logger.info("="*70)
    logger.info("Generation complete!")


if __name__ == "__main__":
    main()