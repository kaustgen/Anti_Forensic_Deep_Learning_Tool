#!/usr/bin/env python3
"""
Generate BOSS dataset with J-UNIWARD steganography.

Uses J-UNIWARD algorithm for state-of-the-art JPEG steganography.
"""

import logging
from pathlib import Path
from openpyxl import Workbook
from tqdm import tqdm
import random
from j_uniward import JUNIWARDEmbedder

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Configuration
BASE_DIR = Path(__file__).parent
EXCEL_OUT = BASE_DIR / 'BOSS_stego_metadata.xlsx'
TARGET_STEGO_IMAGES = 10000  # Number of stego images to generate

# J-UNIWARD embedding rates (bpnzAC - bits per non-zero AC coefficient)
# Standard research values: 0.1, 0.2, 0.4
EMBEDDING_RATES = {
    'low': {
        'ratio': 0.25,
        'bpnzAC': 0.1,
        'description': 'Low embedding rate (0.1 bpnzAC)'
    },
    'medium': {
        'ratio': 0.25,
        'bpnzAC': 0.2,
        'description': 'Medium embedding rate (0.2 bpnzAC)'
    },
    'high': {
        'ratio': 0.25,
        'bpnzAC': 0.3,
        'description': 'High embedding rate (0.3 bpnzAC)'
    },
    'standard': {
        'ratio': 0.25,
        'bpnzAC': 0.4,
        'description': 'BOSS Standard embedding rate (0.4 bpnzAC)'
    }
}


def main():
    # Find cover images
    images_dir = BASE_DIR / 'BOSS1.01_Dataset'
    available_covers = []
    
    if images_dir.exists() and images_dir.is_dir():
        available_covers.extend(sorted(list(images_dir.glob('*.jpg')) + 
                                      list(images_dir.glob('*.jpeg'))))
    
    if not available_covers:
        raise RuntimeError(f"No cover images found in {images_dir}")
    
    logger.info("Found %d cover images", len(available_covers))
    
    # Setup output folder
    stego_folder = BASE_DIR / "BOSS_sten_data"
    stego_folder.mkdir(parents=True, exist_ok=True)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "File Path"
    ws["B1"] = "Stegnography Applied?"
    ws["C1"] = "Payload Category"
    ws["D1"] = "Payload Size (bytes)"
    ws["E1"] = "Payload (bpp AC DCT)"
    ws["F1"] = "Payload (bytes)"
    ws["G1"] = "Payload (bits)"
    ws["H1"] = "Image Dimensions"
    ws["I1"] = "Non-zero AC DCT"
    ws["J1"] = "RGB"  # New column: False for BOSS (grayscale)
    row_count = 2
    
    # Calculate targets for each embedding rate
    rate_targets = {}
    for category, config in EMBEDDING_RATES.items():
        count = int(TARGET_STEGO_IMAGES * config['ratio'])
        rate_targets[category] = {
            'count': count,
            'bpnzAC': config['bpnzAC'],
            'description': config['description'],
            'generated': 0
        }
    
    logger.info("="*70)
    logger.info("J-UNIWARD EMBEDDING CONFIGURATION:")
    for category, target in rate_targets.items():
        logger.info("  %s: %d images (%s)", 
                   category.capitalize(), 
                   target['count'], 
                   target['description'])
    logger.info("  Total stego target: %d", TARGET_STEGO_IMAGES)
    logger.info("  Total clean images: %d", len(available_covers))
    logger.info("="*70)
    
    # Shuffle covers
    random.seed(42)
    shuffled_covers = list(available_covers)
    random.shuffle(shuffled_covers)
    
    # Distribute covers to embedding rates
    cover_assignments = {'low': [], 'medium': [], 'high': [], 'standard' : []}
    cover_idx = 0
    
    for category in ['low', 'medium', 'high', 'standard']:
        target_count = rate_targets[category]['count']
        # Assign covers (reuse if needed)
        for _ in range(target_count):
            cover_assignments[category].append(shuffled_covers[cover_idx % len(shuffled_covers)])
            cover_idx += 1
    
    # Generate stego images by embedding rate
    logger.info("\nGenerating stego images with J-UNIWARD...")
    
    stats_summary = {
        'total_payload_bytes': 0,
        'total_changes': 0,
        'total_nzAC': 0
    }
    
    for category, covers in cover_assignments.items():
        bpnzAC = rate_targets[category]['bpnzAC']
        logger.info("\n--- Generating %s rate stego images (%.1f bpnzAC) ---", 
                   category, bpnzAC)
        
        embedder = JUNIWARDEmbedder(alpha=bpnzAC)
        category_pbar = tqdm(covers, desc=f"{category.capitalize()} rate")
        
        for cover_path in category_pbar:
            # Add original clean image
            ws.cell(row=row_count, column=1, value=str(cover_path.resolve()))
            ws.cell(row=row_count, column=2, value=False)
            ws.cell(row=row_count, column=3, value='N/A')
            ws.cell(row=row_count, column=4, value='N/A')
            ws.cell(row=row_count, column=5, value='N/A')
            ws.cell(row=row_count, column=6, value='N/A')
            ws.cell(row=row_count, column=7, value='N/A')
            ws.cell(row=row_count, column=8, value='N/A')
            ws.cell(row=row_count, column=9, value='N/A')
            ws.cell(row=row_count, column=10, value=False)  # RGB = False (grayscale)
            row_count += 1
            
            # Generate stego image
            stego_filename = f"stego_{cover_path.stem}_{rate_targets[category]['generated']}.jpg"
            stego_path = stego_folder / stego_filename
            
            try:
                # Embed with J-UNIWARD
                stats = embedder.embed(
                    cover_path=str(cover_path),
                    stego_path=str(stego_path),
                    payload_bytes=None  # Use alpha rate
                )
                
                if stego_path.exists():
                    ws.cell(row=row_count, column=1, value=str(stego_path.resolve()))
                    ws.cell(row=row_count, column=2, value=True)
                    ws.cell(row=row_count, column=3, value=category)
                    ws.cell(row=row_count, column=4, value=stats['payload_bytes'])
                    ws.cell(row=row_count, column=5, value=f"{stats['bpnzAC']:.6f}")
                    ws.cell(row=row_count, column=6, value=stats['payload_bytes'])
                    ws.cell(row=row_count, column=7, value=stats['payload_bits'])
                    ws.cell(row=row_count, column=8, value="512×512")
                    ws.cell(row=row_count, column=9, value=stats['nzAC'])
                    ws.cell(row=row_count, column=10, value=False)  # RGB = False (grayscale)
                    row_count += 1
                    
                    rate_targets[category]['generated'] += 1
                    stats_summary['total_payload_bytes'] += stats['payload_bytes']
                    stats_summary['total_changes'] += stats['changes']
                    stats_summary['total_nzAC'] += stats['nzAC']
                else:
                    logger.warning("J-UNIWARD failed for %s (output not created)", cover_path.name)
                    
            except Exception as e:
                logger.error("Error embedding %s: %s", cover_path.name, str(e))
                continue
    
    # Summary statistics
    logger.info("\n" + "="*70)
    logger.info("GENERATION SUMMARY")
    logger.info("="*70)
    logger.info("Total clean images: %d", len(available_covers))
    logger.info("Total stego images by category:")
    total_stego = 0
    for category, target in rate_targets.items():
        success_rate = 100 * target['generated'] / target['count'] if target['count'] > 0 else 0
        logger.info("  %s (%.1f bpnzAC): %d / %d (%.1f%%)", 
                   category.capitalize(),
                   target['bpnzAC'],
                   target['generated'],
                   target['count'],
                   success_rate)
        total_stego += target['generated']
    
    logger.info("Total stego images generated: %d", total_stego)
    
    if total_stego > 0:
        avg_payload = stats_summary['total_payload_bytes'] / total_stego
        avg_changes = stats_summary['total_changes'] / total_stego
        avg_change_rate = 100 * stats_summary['total_changes'] / (total_stego * 262144)
        
        logger.info("\nEmbedding Statistics:")
        logger.info("  Average payload: %.1f bytes", avg_payload)
        logger.info("  Average changes per image: %.1f", avg_changes)
        logger.info("  Average change rate: %.2f%%", avg_change_rate)
    
    logger.info("="*70)
    
    # Save workbook
    logger.info("\nSaving Excel manifest...")
    wb.save(EXCEL_OUT)
    
    # Calculate actual totals from workbook
    total_clean = 0
    total_stego_in_wb = 0
    category_counts = {'low': 0, 'medium': 0, 'high': 0, 'standard': 0}
    
    for row in ws.iter_rows(min_row=2):
        if row[1].value:  # Stego
            total_stego_in_wb += 1
            category = row[2].value
            if category in category_counts:
                category_counts[category] += 1
        else:
            total_clean += 1
    
    # Final summary
    logger.info("\n" + "="*70)
    logger.info("FINAL DATASET SUMMARY")
    logger.info("="*70)
    logger.info("Clean images: %d", total_clean)
    logger.info("Stego images by embedding rate:")
    for category in ['low', 'medium', 'high', 'standard']:
        percentage = 100 * category_counts[category] / total_stego_in_wb if total_stego_in_wb > 0 else 0
        logger.info("  %s (%.1f bpnzAC): %d (%.1f%%)", 
                   category.capitalize(), 
                   rate_targets[category]['bpnzAC'],
                   category_counts[category], 
                   percentage)
    logger.info("Total stego: %d", total_stego_in_wb)
    logger.info("Total images: %d", total_clean + total_stego_in_wb)
    logger.info("Stego/Clean ratio: %.2f", total_stego_in_wb / total_clean if total_clean > 0 else 0)
    logger.info("RGB: False (all grayscale BOSS images)")
    logger.info("="*70)
    logger.info("Dataset manifest saved to: %s", EXCEL_OUT)
    logger.info("Generation complete!")


if __name__ == "__main__":
    main()
